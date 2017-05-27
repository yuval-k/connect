from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import sys

wb = load_workbook(filename=sys.argv[1], data_only=True)
sheet = wb["20Ex"]

strips = []
for pair in range(10):
    strip1 = []
    strip2 = []
    col = 3 + pair*4

    for stripindex in range(54):
        pixelrow = 5 + stripindex
        zcell  = sheet.cell(column=2,     row=pixelrow)
        xcellp = sheet.cell(column=col,   row=pixelrow)
        ycellp = sheet.cell(column=col+1, row=pixelrow)
        xcellm = sheet.cell(column=col+2, row=pixelrow)
        ycellm = sheet.cell(column=col+3, row=pixelrow)

        z = zcell.value / 100.0
        strip1.append((xcellp.value/ 100.0, ycellp.value/ 100.0, z))
        strip2.append((xcellm.value/ 100.0, ycellm.value/ 100.0, z))

    # add heart LEDs:
    lastx,lasty,lastz = strip1[-1]
    strip1 = strip1  + [(lastx,lasty,lastz-.1)]*14

    lastx,lasty,lastz = strip2[-1]
    strip2 = strip2  + [(lastx,lasty,lastz-.1)]*14

    # opc server assumes 100 pixels per strips
    strip1 = strip1 + [(0,0,0)]*(100-len(strip1))
    strip2 = strip2 + [(0,0,0)]*(100-len(strip2))


    strip1.sort(key=lambda x:-x[2])
    strip2.sort(key=lambda x:-x[2])

    strips += [strip1, strip2]

pixels = []
for strip in strips:
    for pixel in strip:
        pixels.append({"point" : pixel})

import json
print json.dumps(pixels)